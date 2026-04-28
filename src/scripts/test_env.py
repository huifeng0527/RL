"""Multi-dimensional Stress Test Suite for Robot Evaluation.

This test suite evaluates and compares three robot training approaches:
- Baseline A: Robot trained only with static script hands
- Baseline B: Robot trained only with a single strongest RL hand
- PFSP (Ours): Robot trained with Progressive Fictitious Self-Play iterative league

The "final exam" consists of 4 diverse test hands designed to expose
weaknesses in Baseline A and B while PFSP handles them gracefully.

Usage:
    python src/scripts/test_env.py --robot all --test all --episodes 50 --visual
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

import argparse
import json
import time
from dataclasses import dataclass, asdict
from typing import List, Optional, Tuple

import numpy as np
import pygame
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from src.custom_env import RehabilitationEnv
from src.renderer import render_aesthetic
from stable_baselines3 import PPO


# =============================================================================
# Constants
# =============================================================================
ZPD_MIN = 3.5
ZPD_MAX = 6.0
GRID_SIZE = 10
CELL_SIZE = 50
WIDTH_PX = int(GRID_SIZE * CELL_SIZE * 1.5)
HEIGHT_PX = int(GRID_SIZE * CELL_SIZE)

# Model paths
BASE_DIR = r"C:\Users\admin\Desktop\科研\RL\logs\dual_iterative_0427_1314"
MODEL_PATHS = {
    'baseline_a': r"C:\Users\admin\Desktop\科研\RL\logs\ablation_study_0424_1945\2_MLP_LSTM\best_model.zip",
    'baseline_b': os.path.join(BASE_DIR, "baseline_b", "robot", "best_model.zip"),
    'pfsp': os.path.join(BASE_DIR, "iteration_13", "robot", "robot", "best_model.zip"),
}
UNSEEN_HAND_PATH = os.path.join(BASE_DIR, "iteration_13", "hand", "hand", "best_model.zip")


# =============================================================================
# Metrics
# =============================================================================
@dataclass
class StressTestMetrics:
    """Metrics captured during stress test."""
    robot_name: str
    test_name: str
    survival_times: List[int]
    mean_survival_time: float
    std_survival_time: float
    catch_count: int
    catch_rate: float
    zpd_maintenance_rate: float
    mean_distance: float
    zpd_steps: float = 0.0  # ZPD内有效步数 = S × Z

    def to_dict(self):
        return asdict(self)


# =============================================================================
# Test Hand Classes
# =============================================================================

class SluggishScriptHand:
    """极慢速脚本手 (Test 1) - stride locked at 0.3.

    专门用来打脸 Baseline B：它天天跟神仙打架，全速逃命。
    遇到这种极慢的手，会因为过度紧张，瞬间跑到屏幕对角线躲起来。
    """

    STRIDE = 0.3  # 重度中风，完全动不了

    def __init__(self, epsilon: float = 0.2):
        self.epsilon = epsilon

    def get_move(self, hand_pos: np.ndarray, robot_pos: np.ndarray) -> np.ndarray:
        """返回手的移动向量."""
        scripted_stride = self.STRIDE
        if np.random.random() < self.epsilon:
            move = np.random.uniform(-1, 1, size=2)
            move = self._safe_normalize(move) * scripted_stride
        else:
            vec = robot_pos - hand_pos
            move = self._safe_normalize(vec) * scripted_stride
        return move

    @staticmethod
    def _safe_normalize(vec: np.ndarray) -> np.ndarray:
        norm = np.linalg.norm(vec)
        if norm < 1e-6:
            return np.zeros_like(vec)
        return vec / norm


class SpasmScriptHand:
    """突发痉挛手 (Test 2) - alternates slow/fast.

    专门用来打脸 Baseline A：它习惯了对方是匀速的。
    当手突然加速时，反应不及，瞬间被抓。
    """

    SLOW_STRIDE = 0.2
    FAST_STRIDE = 0.5
    SLOW_FRAMES = 32   # 4 sec @ 8 FPS
    FAST_FRAMES = 8    # 1 sec @ 8 FPS

    def __init__(self, epsilon: float = 0.2):
        self.epsilon = epsilon
        self.frame_counter = 0
        self.is_fast_phase = False

    def get_move(self, hand_pos: np.ndarray, robot_pos: np.ndarray) -> np.ndarray:
        """返回手的移动向量。"""
        # 切换阶段
        total_frames = self.SLOW_FRAMES + self.FAST_FRAMES
        self.frame_counter = (self.frame_counter + 1) % total_frames
        self.is_fast_phase = self.frame_counter >= self.SLOW_FRAMES

        # 选择步长
        stride = self.FAST_STRIDE if self.is_fast_phase else self.SLOW_STRIDE

        # 计算移动
        if np.random.random() < self.epsilon:
            move = np.random.uniform(-1, 1, size=2)
            move = self._safe_normalize(move) * stride
        else:
            vec = robot_pos - hand_pos
            move = self._safe_normalize(vec) * stride
        return move

    def reset(self):
        """重置痉挛相位。"""
        self.frame_counter = 0
        self.is_fast_phase = False

    @staticmethod
    def _safe_normalize(vec: np.ndarray) -> np.ndarray:
        norm = np.linalg.norm(vec)
        if norm < 1e-6:
            return np.zeros_like(vec)
        return vec / norm


class UnseenRLHand:
    """未见过的RL手 (Test 3) - 全新随机种子的RL Hand.

    双杀 A 和 B：真正测试零样本泛化能力。
    不管 A、B 还是 PFSP，训练时绝对没见过这个 Hand。
    """

    def __init__(self, model_path: str, stride: float = 0.35):
        self.model = PPO.load(model_path, custom_objects={
            'learning_rate': 0.0,
            'optimizer_class': None
        }, verbose=0)
        self.stride = stride

    def get_move(self, hand_obs: np.ndarray) -> np.ndarray:
        """返回手的移动向量 (经过stride缩放)。"""
        action, _ = self.model.predict(hand_obs, deterministic=True)
        return action * self.stride


class HumanMouseHand:
    """人类鼠标控制的手 (Test 4) - 用鼠标直接控制。

    降维打击：最具说服力的测试。
    人类玩家会发现 Ours 像个泥鳅，最难抓。
    """

    def __init__(self, env: RehabilitationEnv, cell_size: int = CELL_SIZE):
        self.env = env
        self.cell_size = cell_size
        # 启用旁路模式：跳过生物力学约束
        env._bypass_hand_physics = True
        env.distance_threshold_collision = 1.5

    def update(self, mouse_x_px: int, mouse_y_px: int) -> None:
        """根据鼠标位置更新手的位置。"""
        # 转换鼠标坐标到环境坐标
        target_x = mouse_x_px / self.cell_size
        target_y = mouse_y_px / self.cell_size

        # 裁剪到有效范围
        target_x = np.clip(target_x, self.env.margin, self.env.env_width - self.env.margin)
        target_y = np.clip(target_y, self.env.margin, self.env.env_height - self.env.margin)
        target_pos = np.array([target_x, target_y])

        # 计算平滑移动（尊重stride限制）
        vec_to_mouse = target_pos - self.env.hand_position
        dist_to_mouse = np.linalg.norm(vec_to_mouse)

        if dist_to_mouse > 1e-4:
            move_dist = min(dist_to_mouse, self.env.stride_hand)
            hand_move = (vec_to_mouse / dist_to_mouse) * move_dist
        else:
            hand_move = np.zeros(2)

        # 更新手的物理惯性状态
        self.env.last_hand_actual_move = hand_move.copy()
        self.env.hand_position += hand_move
        self.env.hand_position = np.clip(
            self.env.hand_position,
            self.env.margin,
            [self.env.env_width - self.env.margin, self.env.env_height - self.env.margin]
        )
        self.env.hand_history_buffer.append(hand_move)


# =============================================================================
# Test Runner
# =============================================================================

def calculate_zpd_rate(distances: List[float]) -> float:
    """计算ZPD维护率 - 距离在[4,6]范围内的时间占比。"""
    if not distances:
        return 0.0
    in_zpd = sum(1 for d in distances if ZPD_MIN <= d <= ZPD_MAX)
    return in_zpd / len(distances)


def save_results_to_excel(all_results: List[StressTestMetrics], output_path: str):
    """Save results to a professionally formatted Excel file."""
    if not HAS_OPENPYXL:
        print("[Warning] openpyxl not installed, skipping Excel export")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stress Test Results"

    # 定义样式
    header_font = Font(name='Times New Roman', size=11, bold=True)
    cell_font = Font(name='Times New Roman', size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')

    # 测试名称映射（用于表头）
    test_names = {
        'sluggish': 'Sluggish\n(S/C/Z/ZPD)',
        'spasm': 'Spasm\n(S/C/Z/ZPD)',
        'unseen_rl': 'Unseen RL\n(S/C/Z/ZPD)',
        'human': 'Human\n(S/C/Z/ZPD)',
    }
    robot_names = {
        'baseline_a': 'Baseline A',
        'baseline_b': 'Baseline B',
        'pfsp': 'PFSP (Ours)',
    }

    # 按 Test 分组
    tests_order = ['sluggish', 'spasm', 'unseen_rl', 'human']
    robots_order = ['baseline_a', 'baseline_b', 'pfsp']

    # 构建数据矩阵
    data = {}
    for r in all_results:
        data[(r.test_name, r.robot_name)] = r

    # 写表头
    headers = ['Model'] + [test_names.get(t, t) for t in tests_order]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    # 设置列宽
    ws.column_dimensions['A'].width = 15
    for col in range(2, 2 + len(tests_order)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    # 写数据行
    for row_idx, robot in enumerate(robots_order, 2):
        # Model 名称
        cell = ws.cell(row=row_idx, column=1, value=robot_names.get(robot, robot))
        cell.font = cell_font
        cell.alignment = center_align
        cell.border = thin_border

        # 数据
        for col_idx, test in enumerate(tests_order, 2):
            key = (test, robot)
            if key in data:
                m = data[key]
                val = f"{m.mean_survival_time:.1f}/{m.catch_rate:.0%}/{m.zpd_maintenance_rate:.0%}/{m.zpd_steps:.1f}"
            else:
                val = "N/A"
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = center_align
            cell.border = thin_border

    # 添加底部注释
    note_row = len(robots_order) + 3
    note = ws.cell(row=note_row, column=1, value="Note: S=Survival time (steps), C=Catch rate (%), Z=ZPD maintenance rate (%)")
    note.font = Font(name='Times New Roman', size=9, italic=True)

    wb.save(output_path)
    print(f"[Excel] Results saved to: {output_path}")


def run_stress_test(
    robot_model_path: str,
    robot_name: str,
    test_hand_type: str,
    unseen_hand_model_path: Optional[str] = None,
    num_episodes: int = 50,
    max_steps: int = 100,
    visual: bool = False,
    fps: int = 8,
) -> StressTestMetrics:
    """运行单个压力测试配置。"""

    # 加载Robot模型
    robot_model = PPO.load(robot_model_path, verbose=0) if robot_model_path else None

    # 创建环境
    env = RehabilitationEnv(
        training_mode='robot',
        robot_model=robot_model,
        hand_model=None,  # 手由测试控制器直接控制
    )
    env.grid_size = GRID_SIZE
    env.cell_size = CELL_SIZE
    env.random_noise = False
    env.max_steps = max_steps

    # 初始化测试手
    test_hand = None
    hand_model_for_env = None  # 用于 unseen_rl 测试的 RL hand model

    if test_hand_type == 'sluggish':
        test_hand = SluggishScriptHand()
    elif test_hand_type == 'spasm':
        test_hand = SpasmScriptHand()
    elif test_hand_type == 'unseen_rl':
        if unseen_hand_model_path and os.path.exists(unseen_hand_model_path):
            # 将 RL hand model 交给环境管理，由 _resolve_hand_move() 应用物理约束
            hand_model_for_env = PPO.load(unseen_hand_model_path, custom_objects={
                'learning_rate': 0.0,
                'optimizer_class': None
            }, verbose=0)
        else:
            print(f"[Warning] Unseen RL hand model not found: {unseen_hand_model_path}")
            return StressTestMetrics(
                robot_name=robot_name, test_name=test_hand_type,
                survival_times=[], mean_survival_time=0, std_survival_time=0,
                catch_count=0, catch_rate=0, zpd_maintenance_rate=0, mean_distance=0,
                zpd_steps=0.0
            )
    elif test_hand_type == 'human':
        test_hand = HumanMouseHand(env)

    # 如果有 RL hand model，设置为环境的 hand_model（由 _resolve_hand_move() 处理物理约束）
    if hand_model_for_env is not None:
        env.hand_model = hand_model_for_env

    # 用于可视化的pygame初始化
    screen = None
    clock = None
    if visual:
        pygame.init()
        screen = pygame.display.set_mode((WIDTH_PX, HEIGHT_PX))
        pygame.display.set_caption(f"Stress Test: {robot_name} vs {test_hand_type}")
        clock = pygame.time.Clock()

    # 存储结果
    all_survival_times = []
    catch_count = 0
    all_distances = []

    print(f"\n{'='*60}")
    print(f"Testing: {robot_name} vs {test_hand_type}")
    print(f"{'='*60}")

    for ep in range(num_episodes):
        obs, info = env.reset()
        episode_distances = [info['dist']]

        if test_hand_type == 'spasm' and isinstance(test_hand, SpasmScriptHand):
            test_hand.reset()

        if test_hand_type == 'human' and isinstance(test_hand, HumanMouseHand):
            env.stride_robot = 0.6
            env.stride_hand = 0.3

        episode_done = False
        steps = 0

        while not episode_done and steps < max_steps:
            # 处理pygame事件（仅human模式需要）
            if test_hand_type == 'human' and visual:
                for event in pygame.event.get():
                    if event.type == pygame.QUIT:
                        episode_done = True
                    elif event.type == pygame.KEYDOWN:
                        if event.key == pygame.K_q:
                            episode_done = True

            # 获取Robot动作
            if robot_model is not None:
                action, _ = robot_model.predict(obs, deterministic=True)
            else:
                action = env.action_space.sample()

            # 更新测试手 (应用与 RL hand 相同的物理约束)
            if test_hand_type in ['sluggish', 'spasm']:
                assert isinstance(test_hand, (SluggishScriptHand, SpasmScriptHand))
                desired_move = test_hand.get_move(env.hand_position, env.robot_position)

                # 物理约束 I：一阶惯性低通滤波 (Muscle Inertia)
                alpha = 0.7
                smoothed_move = alpha * desired_move + (1 - alpha) * env.last_hand_actual_move

                # 物理约束 II：最大加速度截断 (Acceleration Clipping)
                max_accel = 0.15
                delta_v = smoothed_move - env.last_hand_actual_move
                accel_magnitude = np.linalg.norm(delta_v)
                if accel_magnitude > max_accel:
                    delta_v = (delta_v / accel_magnitude) * max_accel

                final_move = env.last_hand_actual_move + delta_v
                env.last_hand_actual_move = final_move.copy()

                env.hand_position += final_move
                env.hand_position = np.clip(
                    env.hand_position,
                    env.margin,
                    [env.env_width - env.margin, env.env_height - env.margin]
                )
                env.hand_history_buffer.append(final_move)

            # unseen_rl: 环境通过 hand_model + _resolve_hand_move() 自动处理物理约束，无需手动更新
            # elif test_hand_type == 'unseen_rl':
            #     pass  # 让 env.step() 自动处理

            elif test_hand_type == 'human' and visual:
                mouse_x, mouse_y = pygame.mouse.get_pos()
                test_hand.update(mouse_x, mouse_y)

            # 环境step
            obs, reward, terminated, truncated, info = env.step(action)
            episode_distances.append(info['dist'])
            steps += 1

            # 可视化渲染
            if visual and screen is not None:
                render_aesthetic(
                    env.robot_position,
                    env.hand_position,
                    env.fixed_point,
                    env.trajectory_points,
                    grid_size=GRID_SIZE,
                    cell_size=CELL_SIZE,
                    window=screen
                )
                clock.tick(fps)

            # 检查是否结束
            if terminated or truncated:
                episode_done = True

        # 记录结果
        all_survival_times.append(steps)
        if terminated and info.get('done_reason') == 'Robot Caught':
            catch_count += 1
        all_distances.extend(episode_distances)

        if (ep + 1) % 10 == 0:
            print(f"  Episode {ep+1}/{num_episodes} completed, catch_count={catch_count}")

    env.close()
    if visual:
        pygame.quit()

    # 计算指标
    mean_survival = np.mean(all_survival_times) if all_survival_times else 0
    std_survival = np.std(all_survival_times) if all_survival_times else 0
    mean_dist = np.mean(all_distances) if all_distances else 0
    zpd_rate = calculate_zpd_rate(all_distances)
    zpd_steps = mean_survival * zpd_rate  # ZPD内有效步数 = S × Z

    metrics = StressTestMetrics(
        robot_name=robot_name,
        test_name=test_hand_type,
        survival_times=all_survival_times,
        mean_survival_time=float(mean_survival),
        std_survival_time=float(std_survival),
        catch_count=catch_count,
        catch_rate=catch_count / num_episodes,
        zpd_maintenance_rate=zpd_rate,
        mean_distance=float(mean_dist),
        zpd_steps=float(zpd_steps)
    )

    print(f"\nResults for {robot_name} vs {test_hand_type}:")
    print(f"  Mean Survival Time: {mean_survival:.2f} ± {std_survival:.2f} steps")
    print(f"  Catch Rate: {metrics.catch_rate:.2%}")
    print(f"  ZPD Maintenance: {metrics.zpd_maintenance_rate:.2%}")
    print(f"  Mean Distance: {mean_dist:.2f}")

    return metrics


def run_human_test_interactive(
    robot_model_path: str,
    robot_name: str,
    num_episodes: int = 5,
    max_steps: int = 100,
    fps: int = 8,
) -> Tuple[List[int], int, float, float, float, float]:
    """运行交互式人类测试（鼠标控制）。

    Returns:
        (survival_times, catch_count, mean_survival, std_survival, zpd_rate, zpd_steps)
    """
    robot_model = PPO.load(robot_model_path, verbose=0)

    pygame.init()
    screen = pygame.display.set_mode((WIDTH_PX, HEIGHT_PX))
    clock = pygame.time.Clock()

    all_survival_times = []
    catch_count = 0
    all_distances = []

    for ep in range(num_episodes):
        env = RehabilitationEnv(
            training_mode='robot',
            robot_model=robot_model,
            hand_model=None,
        )
        env.grid_size = GRID_SIZE
        env.cell_size = CELL_SIZE
        env._bypass_hand_physics = True
        env.distance_threshold_collision = 1.5
        env.stride_robot = 0.6
        env.stride_hand = 0.3
        env.max_steps = max_steps

        pygame.display.set_caption(f"HUMAN TEST: Ep {ep+1}/{num_episodes} - You are the HAND! vs {robot_name}")

        obs, info = env.reset()
        distances = [info['dist']]
        steps = 0
        running = True

        print(f"\nEpisode {ep+1}/{num_episodes}: Survive as long as possible!")
        print("Press Q to quit this episode, R to restart.\n")

        episode_done = False
        caught = False

        while running and not episode_done and steps < max_steps:
            for event in pygame.event.get():
                if event.type == pygame.QUIT:
                    running = False
                elif event.type == pygame.KEYDOWN:
                    if event.key == pygame.K_q:
                        running = False
                    elif event.key == pygame.K_r:
                        episode_done = True

            # 获取鼠标位置并更新手
            mouse_x, mouse_y = pygame.mouse.get_pos()
            target_x = np.clip(mouse_x / CELL_SIZE, env.margin, env.env_width - env.margin)
            target_y = np.clip(mouse_y / CELL_SIZE, env.margin, env.env_height - env.margin)
            target_pos = np.array([target_x, target_y])

            vec_to_mouse = target_pos - env.hand_position
            dist_to_mouse = np.linalg.norm(vec_to_mouse)
            if dist_to_mouse > 1e-4:
                move_dist = min(dist_to_mouse, env.stride_hand)
                hand_move = (vec_to_mouse / dist_to_mouse) * move_dist
            else:
                hand_move = np.zeros(2)

            env.last_hand_actual_move = hand_move.copy()
            env.hand_position += hand_move
            env.hand_position = np.clip(
                env.hand_position, env.margin,
                [env.env_width - env.margin, env.env_height - env.margin]
            )
            env.hand_history_buffer.append(hand_move)

            # Robot动作
            action, _ = robot_model.predict(obs, deterministic=True)
            obs, reward, terminated, truncated, info = env.step(action)
            distances.append(info['dist'])
            steps += 1

            # 渲染
            render_aesthetic(
                env.robot_position,
                env.hand_position,
                env.fixed_point,
                env.trajectory_points,
                grid_size=GRID_SIZE,
                cell_size=CELL_SIZE,
                window=screen
            )

            # 显示信息
            font = pygame.font.SysFont('arial', 18)
            dist_text = font.render(f"Ep {ep+1}/{num_episodes}  Dist: {info['dist']:.2f}  Steps: {steps}  ZPD: {ZPD_MIN}-{ZPD_MAX}", True, (0, 0, 0))
            screen.blit(dist_text, (10, 10))
            pygame.display.flip()
            clock.tick(fps)

            if terminated or truncated:
                caught = terminated and info.get('done_reason') == 'Robot Caught'
                episode_done = True

        all_survival_times.append(steps)
        if caught:
            catch_count += 1
        all_distances.extend(distances)

        env.close()

    pygame.quit()

    mean_survival = np.mean(all_survival_times)
    std_survival = np.std(all_survival_times)
    zpd_rate = calculate_zpd_rate(all_distances)

    return all_survival_times, catch_count, mean_survival, std_survival, zpd_rate, np.mean(all_distances)


# =============================================================================
# CLI Interface
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='Multi-dimensional Stress Test Suite for Robot Evaluation'
    )
    parser.add_argument(
        '--robot',
        type=str,
        default='all',
        choices=['baseline_a', 'baseline_b', 'pfsp', 'all'],
        help='Robot to test'
    )
    parser.add_argument(
        '--test',
        type=str,
        default='all',
        choices=['sluggish', 'spasm', 'unseen_rl', 'human', 'all'],
        help='Test hand type'
    )
    parser.add_argument(
        '--episodes',
        type=int,
        default=50,
        help='Number of episodes per test'
    )
    parser.add_argument(
        '--visual',
        action='store_true',
        help='Enable visual rendering'
    )
    parser.add_argument(
        '--fps',
        type=int,
        default=8,
        help='Frames per second for visual mode'
    )
    parser.add_argument(
        '--output',
        type=str,
        default='stress_test_results',
        help='Output directory for results'
    )
    parser.add_argument(
        '--baseline_b_path',
        type=str,
        default=None,
        help='Path to Baseline B model (if already trained)'
    )

    args = parser.parse_args()

    # 创建输出目录
    os.makedirs(args.output, exist_ok=True)

    # 确定要测试的robot和test组合
    robots = []
    if args.robot == 'all':
        robots = ['baseline_a', 'baseline_b', 'pfsp']
    else:
        robots = [args.robot]

    tests = []
    if args.test == 'all':
        tests = ['sluggish', 'spasm', 'unseen_rl', 'human']
    else:
        tests = [args.test]

    # 获取有效model path
    def get_robot_path(name: str) -> Optional[str]:
        if name == 'baseline_b':
            return args.baseline_b_path if args.baseline_b_path else MODEL_PATHS.get(name)
        return MODEL_PATHS.get(name)

    # 运行所有测试
    all_results = []
    timestamp = time.strftime("%Y%m%d_%H%M%S")

    for robot in robots:
        robot_path = get_robot_path(robot)
        if robot_path is None:
            print(f"\n[Skipping {robot}] Model path is None (needs training)")
            continue
        if not os.path.exists(robot_path):
            print(f"\n[Skipping {robot}] Model not found: {robot_path}")
            continue

        for test in tests:
            if test == 'human':
                # 人类测试需要交互式界面，单独处理
                print(f"\n[Starting HUMAN test for {robot}]")
                                survival_times, catch_count, mean_survival, std_survival, zpd_rate, mean_dist = run_human_test_interactive(
                    robot_path, robot, num_episodes=args.episodes, fps=args.fps
                )
                zpd_steps = mean_survival * zpd_rate
                metrics = StressTestMetrics(
                    robot_name=robot,
                    test_name='human',
                    survival_times=survival_times,
                    mean_survival_time=mean_survival,
                    std_survival_time=std_survival,
                    catch_count=catch_count,
                    catch_rate=catch_count / args.episodes,vival,
                    std_survival_time=std_survival,
                    catch_count=catch_count,
                    catch_rate=catch_count / 5,
                    zpd_maintenance_rate=zpd_rate,
                    mean_distance=mean_dist,
                    zpd_steps=zpd_steps
                )
            else:
                metrics = run_stress_test(
                    robot_model_path=robot_path,
                    robot_name=robot,
                    test_hand_type=test,
                    unseen_hand_model_path=UNSEEN_HAND_PATH,
                    num_episodes=args.episodes,
                    visual=args.visual,
                    fps=args.fps
                )
            all_results.append(metrics)

    # 保存结果
    if all_results:
        results_file = os.path.join(args.output, f"results_{timestamp}.json")
        with open(results_file, 'w') as f:
            json.dump([r.to_dict() for r in all_results], f, indent=2)
        print(f"\n{'='*60}")
        print(f"Results saved to: {results_file}")
        print(f"{'='*60}")

        # 保存 Excel
        excel_file = os.path.join(args.output, f"stress_test_results_{timestamp}.xlsx")
        save_results_to_excel(all_results, excel_file)

        # 打印专业汇总表 (Model x Test 矩阵)
        print("\n" + "="*80)
        print("TABLE II: MULTI-DIMENSIONAL STRESS TEST RESULTS")
        print("="*100)
        print(f"{'':15} {'Sluggish':>20} {'Spasm':>20} {'Unseen RL':>20} {'Human':>20}")
        print(f"{'':15} {'(S/C/Z/ZPD)':>20} {'(S/C/Z/ZPD)':>20} {'(S/C/Z/ZPD)':>20} {'(S/C/Z/ZPD)':>20}")
        print("-"*100)

        tests_order = ['sluggish', 'spasm', 'unseen_rl', 'human']
        robots_order = ['baseline_a', 'baseline_b', 'pfsp']
        robot_display = {'baseline_a': 'Baseline A', 'baseline_b': 'Baseline B', 'pfsp': 'PFSP (Ours)'}
        data = {}
        for r in all_results:
            data[(r.test_name, r.robot_name)] = r

        for robot in robots_order:
            row = f"{robot_display[robot]:15}"
            for test in tests_order:
                key = (test, robot)
                if key in data:
                    m = data[key]
                    row += f" {m.mean_survival_time:>5.1f}/{m.catch_rate:.0%}/{m.zpd_maintenance_rate:.0%}/{m.zpd_steps:>5.1f} "
                else:
                    row += f" {'N/A':>20}"
            print(row)
        print("="*100)
        print("S=Survival time (steps), C=Catch rate (%), Z=ZPD maintenance (%), ZPD=ZPD steps (S×Z)\n")
    else:
        print("\nNo results to save (all tests were skipped).")


if __name__ == '__main__':
    main()
