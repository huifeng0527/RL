"""Export raw ablation metrics to Excel.

The Excel workbook stores raw per-episode data. Figure smoothing is controlled
only by plot_ablation_from_excel.py --smooth_window.

Example:
    python src/scripts/export_ablation_review.py --run_dir logs/ablation_gru_h1_h10_0626_2136
"""

import argparse
import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook


GROUP_LABELS = {
    "1_MLP": "MLP",
    "2_GRU_Seq": "GRU",
    "3_GRU_Aux": "GRU + Aux",
}

EPISODE_RAW_COLUMNS = [
    "group",
    "label",
    "episode_index",
    "timestep",
    "timestep_m",
    "reward",
    "episode_length",
    "zpd_steps",
    "zpd_coverage",
    "workspace_coverage",
]

MAX_EXCEL_ROWS_PER_SHEET = 1_000_000


def load_config(run_dir):
    path = run_dir / "config.json"
    if not path.exists():
        return {}
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_episode_metrics(run_dir, groups):
    data = {}
    for group in groups:
        path = run_dir / group / "episode_metrics.npz"
        if not path.exists():
            continue
        npz = np.load(path)
        record = {k: npz[k] for k in npz.files}
        if "zpd_steps" not in record:
            record["zpd_steps"] = record["zpd_coverages"].astype(float) * record["ep_lengths"].astype(float)
        data[group] = record
    return data


def load_eval_metrics(run_dir, groups):
    rows = []
    for group in groups:
        path = run_dir / group / "ablation_metrics.json"
        if not path.exists():
            continue
        with open(path, "r", encoding="utf-8") as f:
            records = json.load(f)
        for item in records:
            row = {"group": group, "label": GROUP_LABELS.get(group, group)}
            row.update(item)
            row["timestep_m"] = row["timestep"] / 1e6
            rows.append(row)
    return pd.DataFrame(rows)


def summarize_episode_metrics(data, tail_fraction=0.1, last_n=1000):
    rows = []
    metrics = {
        "reward": "rewards",
        "episode_length": "ep_lengths",
        "zpd_steps": "zpd_steps",
        "zpd_coverage": "zpd_coverages",
        "workspace_coverage": "workspace_areas",
    }
    for group, record in data.items():
        n = len(record["timesteps"])
        tail_start = int(n * (1.0 - tail_fraction))
        last_n_start = max(0, n - last_n)
        row = {
            "group": group,
            "label": GROUP_LABELS.get(group, group),
            "episodes": n,
            "last_timestep": int(record["timesteps"][-1]) if n else 0,
            "last_timestep_m": float(record["timesteps"][-1] / 1e6) if n else 0.0,
        }
        for out_name, key in metrics.items():
            values = record[key].astype(float)
            row[f"{out_name}_all_mean"] = float(np.mean(values))
            row[f"{out_name}_all_std"] = float(np.std(values))
            row[f"{out_name}_last10_mean"] = float(np.mean(values[tail_start:]))
            row[f"{out_name}_last10_std"] = float(np.std(values[tail_start:]))
            row[f"{out_name}_last{last_n}_mean"] = float(np.mean(values[last_n_start:]))
            row[f"{out_name}_last{last_n}_std"] = float(np.std(values[last_n_start:]))
        rows.append(row)
    return pd.DataFrame(rows)


def summarize_eval_metrics(eval_df):
    if eval_df.empty:
        return pd.DataFrame()
    rows = []
    for group, sub in eval_df.groupby("group", sort=False):
        sub = sub.sort_values("timestep")
        last = sub.iloc[-1]
        best_reward = sub.loc[sub["reward_mean"].idxmax()]
        best_zpd = sub.loc[sub["zpd_coverage_mean"].idxmax()]
        rows.append({
            "group": group,
            "label": GROUP_LABELS.get(group, group),
            "last_timestep": int(last["timestep"]),
            "last_reward_mean": float(last["reward_mean"]),
            "last_episode_length_mean": float(last["episode_length_mean"]),
            "last_zpd_coverage_mean": float(last["zpd_coverage_mean"]),
            "best_reward_timestep": int(best_reward["timestep"]),
            "best_reward_mean": float(best_reward["reward_mean"]),
            "best_reward_episode_length_mean": float(best_reward["episode_length_mean"]),
            "best_reward_zpd_coverage_mean": float(best_reward["zpd_coverage_mean"]),
            "best_zpd_timestep": int(best_zpd["timestep"]),
            "best_zpd_reward_mean": float(best_zpd["reward_mean"]),
            "best_zpd_coverage_mean": float(best_zpd["zpd_coverage_mean"]),
        })
    return pd.DataFrame(rows)


def append_dataframe_sheet(wb, sheet_name, df):
    ws = wb.create_sheet(sheet_name)
    if df.empty:
        ws.append(["empty"])
        return
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))


def append_config_sheet(wb, config):
    ws = wb.create_sheet("Config")
    ws.append(["key", "value"])
    for key, value in config.items():
        if isinstance(value, (list, dict)):
            value = json.dumps(value, ensure_ascii=False)
        ws.append([key, value])


def append_readme_sheet(wb, run_dir):
    ws = wb.create_sheet("README")
    ws.append(["item", "value"])
    ws.append(["Run directory", str(run_dir)])
    ws.append(["Data policy", "Episode_Raw sheets contain raw per-episode data. No smoothing is applied in Excel."])
    ws.append(["Plot control", "Use plot_ablation_from_excel.py --smooth_window to control smoothing."])
    ws.append(["Raw columns", ", ".join(EPISODE_RAW_COLUMNS)])


def append_raw_episode_sheets(wb, data):
    sheet_idx = 1
    ws = wb.create_sheet(f"Episode_Raw_{sheet_idx}")
    ws.append(EPISODE_RAW_COLUMNS)
    rows_in_sheet = 1

    for group, record in data.items():
        n = len(record["timesteps"])
        label = GROUP_LABELS.get(group, group)
        timesteps = record["timesteps"]
        rewards = record["rewards"]
        lengths = record["ep_lengths"]
        zpd_steps = record["zpd_steps"]
        zpd_coverages = record["zpd_coverages"]
        workspace = record["workspace_areas"]

        for i in range(n):
            if rows_in_sheet >= MAX_EXCEL_ROWS_PER_SHEET:
                sheet_idx += 1
                ws = wb.create_sheet(f"Episode_Raw_{sheet_idx}")
                ws.append(EPISODE_RAW_COLUMNS)
                rows_in_sheet = 1
            ws.append([
                group,
                label,
                i + 1,
                int(timesteps[i]),
                float(timesteps[i] / 1e6),
                float(rewards[i]),
                int(lengths[i]),
                float(zpd_steps[i]),
                float(zpd_coverages[i]),
                float(workspace[i]),
            ])
            rows_in_sheet += 1


def write_excel(run_dir, config, data, episode_summary, eval_summary, eval_df, output_name):
    out_path = run_dir / output_name
    wb = Workbook(write_only=True)
    append_readme_sheet(wb, run_dir)
    append_config_sheet(wb, config)
    append_dataframe_sheet(wb, "Episode_Summary", episode_summary)
    append_dataframe_sheet(wb, "Eval_Summary", eval_summary)
    append_dataframe_sheet(wb, "Eval_All", eval_df)
    append_raw_episode_sheets(wb, data)
    wb.save(out_path)
    return out_path


def main():
    parser = argparse.ArgumentParser(description="Export raw ablation metrics to Excel.")
    parser.add_argument("--run_dir", default="logs/ablation_gru_h1_h10_0626_2136")
    parser.add_argument("--output", default="ablation_review_raw.xlsx")
    parser.add_argument("--tail_fraction", type=float, default=0.1)
    parser.add_argument("--last_n", type=int, default=1000)
    args = parser.parse_args()

    run_dir = Path(args.run_dir)
    config = load_config(run_dir)
    groups = config.get("experiments") or ["1_MLP", "2_GRU_Seq", "3_GRU_Aux"]
    data = load_episode_metrics(run_dir, groups)
    if not data:
        raise FileNotFoundError(f"No episode_metrics.npz found under {run_dir}")

    eval_df = load_eval_metrics(run_dir, list(data.keys()))
    episode_summary = summarize_episode_metrics(data, tail_fraction=args.tail_fraction, last_n=args.last_n)
    eval_summary = summarize_eval_metrics(eval_df)

    excel_path = write_excel(run_dir, config, data, episode_summary, eval_summary, eval_df, args.output)
    episode_summary.to_csv(run_dir / "summary_for_paper.csv", index=False)
    print(f"Excel raw data: {excel_path}")
    print(f"Summary CSV: {run_dir / 'summary_for_paper.csv'}")


if __name__ == "__main__":
    main()
